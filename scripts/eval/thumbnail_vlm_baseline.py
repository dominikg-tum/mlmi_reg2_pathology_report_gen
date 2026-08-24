#!/usr/bin/env python
from __future__ import annotations

import argparse
import base64
import json
import mimetypes
import re
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openai import OpenAI
from openpyxl import Workbook, load_workbook


LABELS = [
    "normal_or_no_significant_abnormality",
    "benign_tumor",
    "inflammatory_or_reactive",
    "malignant_tumor",
    "insufficient_or_uncertain",
    "precancerous_lesion",
]


PROMPT = f"""You are classifying a pathology whole-slide thumbnail.

Return exactly one class label from this allowed list:
{", ".join(LABELS)}

Use the visual evidence in the thumbnail only. If the thumbnail is too low resolution,
ambiguous, or does not contain enough evidence, return insufficient_or_uncertain.

Output format:
{{"classification": "<one allowed label>"}}
"""


def image_url(path: Path) -> str:
    mime = mimetypes.guess_type(path.name)[0] or "image/jpeg"
    data = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{data}"


def collect_images(paths: list[Path]) -> list[Path]:
    images: list[Path] = []
    for root in paths:
        if root.is_file():
            images.append(root)
        elif root.is_dir():
            images.extend(
                p
                for p in root.iterdir()
                if p.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}
            )
    return sorted(set(images), key=lambda p: p.name)


def parse_label(text: str) -> str:
    try:
        obj = json.loads(text)
        value = str(obj.get("classification", "")).strip()
        if value in LABELS:
            return value
    except Exception:
        pass
    for label in LABELS:
        if re.search(rf"\b{re.escape(label)}\b", text):
            return label
    return "insufficient_or_uncertain"


def load_done(output: Path) -> set[str]:
    if not output.exists():
        return set()
    wb = load_workbook(output)
    ws = wb.active
    done: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            done.add(str(row[0]))
    return done


def open_workbook(output: Path) -> tuple[Workbook, Any]:
    if output.exists():
        wb = load_workbook(output)
        return wb, wb.active
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "thumbnail_baseline"
    ws.append(["wsi_name", "classification", "thumbnail_path", "raw_response", "status", "created_at"])
    wb.save(output)
    return wb, ws


def classify(client: OpenAI, model: str, image_path: Path, temperature: float, max_tokens: int) -> tuple[str, str]:
    response = client.chat.completions.create(
        model=model,
        temperature=temperature,
        max_tokens=max_tokens,
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": PROMPT},
                    {"type": "image_url", "image_url": {"url": image_url(image_path)}},
                ],
            }
        ],
    )
    raw = response.choices[0].message.content or ""
    return parse_label(raw), raw


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--image-dir", type=Path, action="append", required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--endpoint", default="http://127.0.0.1:8010/v1")
    parser.add_argument("--model", default="/mnt/research/ljs/Qwen3-VL-8B-Instruct")
    parser.add_argument("--api-key", default="token-abc123")
    parser.add_argument("--limit", type=int, default=0, help="0 means all thumbnails")
    parser.add_argument("--sleep", type=float, default=0.0)
    parser.add_argument("--temperature", type=float, default=0.0)
    parser.add_argument("--max-tokens", type=int, default=64)
    parser.add_argument("--retries", type=int, default=3)
    args = parser.parse_args()

    images = collect_images(args.image_dir)
    if args.limit:
        images = images[: args.limit]
    if not images:
        raise SystemExit("No thumbnail images found.")

    done = load_done(args.output)
    wb, ws = open_workbook(args.output)
    client = OpenAI(base_url=args.endpoint, api_key=args.api_key)

    print(f"Found {len(images)} thumbnails; already done {len(done)}; output={args.output}", flush=True)
    for idx, image_path in enumerate(images, start=1):
        wsi_name = image_path.stem
        if wsi_name in done:
            continue
        status = "ok"
        raw = ""
        label = "insufficient_or_uncertain"
        for attempt in range(1, args.retries + 1):
            try:
                label, raw = classify(client, args.model, image_path, args.temperature, args.max_tokens)
                break
            except Exception as exc:
                status = f"error_attempt_{attempt}: {type(exc).__name__}: {exc}"
                raw = status
                print(f"[{idx}/{len(images)}] {wsi_name} {status}", flush=True)
                time.sleep(min(30, 2**attempt))
        ws.append(
            [
                wsi_name,
                label,
                str(image_path),
                raw,
                status if status.startswith("error_attempt_") else "ok",
                datetime.now(timezone.utc).isoformat(),
            ]
        )
        wb.save(args.output)
        done.add(wsi_name)
        print(f"[{idx}/{len(images)}] {wsi_name} -> {label}", flush=True)
        if args.sleep:
            time.sleep(args.sleep)

    wb.save(args.output)
    print(f"Done: {args.output}", flush=True)


if __name__ == "__main__":
    main()
