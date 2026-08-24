#!/usr/bin/env python
from __future__ import annotations

import argparse
import json
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from agent.frontend import run_remote_image_chain


IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp"}


def collect_images(paths: list[Path]) -> list[Path]:
    images: list[Path] = []
    for root in paths:
        if root.is_file() and root.suffix.lower() in IMAGE_SUFFIXES:
            images.append(root)
        elif root.is_dir():
            images.extend(p for p in root.iterdir() if p.suffix.lower() in IMAGE_SUFFIXES)
    return sorted(set(images), key=lambda p: p.name)


def open_workbook(output: Path) -> tuple[Workbook, Any]:
    if output.exists():
        wb = load_workbook(output)
        return wb, wb.active
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "thumbnail_agent_prompt"
    ws.append(
        [
            "wsi_name",
            "vlm_result",
            "status",
            "thumbnail_path",
            "node_path",
            "n_steps",
            "chain_json_path",
            "error",
            "created_at",
        ]
    )
    wb.save(output)
    return wb, ws


def load_done(output: Path) -> set[str]:
    if not output.exists():
        return set()
    wb = load_workbook(output, read_only=True)
    ws = wb.active
    done: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            done.add(str(row[0]))
    return done


def save_chain(chain: dict[str, Any], chains_dir: Path, wsi_name: str) -> Path:
    chains_dir.mkdir(parents=True, exist_ok=True)
    path = chains_dir / f"{wsi_name}.json"
    path.write_text(json.dumps(chain, indent=2, ensure_ascii=False) + "\n")
    return path


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--image-dir", type=Path, action="append", required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--chains-dir", type=Path, required=True)
    parser.add_argument("--endpoint", default="http://127.0.0.1:8000/v1")
    parser.add_argument("--model", default="/mnt/research/ljs/Qwen3-VL-8B-Instruct")
    parser.add_argument("--api-key", default="token-abc123")
    parser.add_argument("--limit", type=int, default=0, help="0 means all thumbnails")
    parser.add_argument("--sleep", type=float, default=0.0)
    parser.add_argument("--retries", type=int, default=2)
    args = parser.parse_args()

    images = collect_images(args.image_dir)
    if args.limit:
        images = images[: args.limit]
    if not images:
        raise SystemExit("No thumbnail images found.")

    done = load_done(args.output)
    wb, ws = open_workbook(args.output)

    print(
        f"Found {len(images)} thumbnails; already done {len(done)}; output={args.output}",
        flush=True,
    )
    for idx, image_path in enumerate(images, start=1):
        wsi_name = image_path.stem
        if wsi_name in done:
            continue

        status = "ok"
        error = ""
        chain: dict[str, Any] = {}
        for attempt in range(1, args.retries + 1):
            try:
                chain = run_remote_image_chain(
                    image_path,
                    base_url=args.endpoint,
                    model_name=args.model,
                    api_key=args.api_key,
                )
                break
            except Exception as exc:
                status = "error"
                error = f"attempt_{attempt}: {type(exc).__name__}: {exc}"
                print(f"[{idx}/{len(images)}] {wsi_name} {error}", flush=True)
                time.sleep(min(30, 2**attempt))

        report = ""
        node_path: list[str] = []
        n_steps = 0
        chain_path = ""
        if chain:
            report = str(chain.get("report") or "")
            node_path = list(chain.get("node_path") or [])
            n_steps = len(chain.get("chain-of-thought") or [])
            chain_path = str(save_chain(chain, args.chains_dir, wsi_name))
            status = "ok"
        else:
            report = ""

        ws.append(
            [
                wsi_name,
                report,
                status,
                str(image_path),
                " -> ".join(node_path),
                n_steps,
                chain_path,
                error,
                datetime.now(timezone.utc).isoformat(),
            ]
        )
        wb.save(args.output)
        done.add(wsi_name)
        print(f"[{idx}/{len(images)}] {wsi_name} -> {status}: {report[:160]}", flush=True)
        if args.sleep:
            time.sleep(args.sleep)

    wb.save(args.output)
    print(f"Done: {args.output}", flush=True)


if __name__ == "__main__":
    main()
