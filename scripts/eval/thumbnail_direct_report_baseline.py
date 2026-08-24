#!/usr/bin/env python
from __future__ import annotations

import argparse
import base64
import json
import mimetypes
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openai import OpenAI
from openpyxl import Workbook, load_workbook

from agent.backends import SYSTEM_PROMPT


IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp"}


DIRECT_REPORT_PROMPT = """Visual evidence: whole-slide thumbnail attached.

Current question:
Generate the final structured pathology report.

Diagnostic guidance:
Synthesize the visible thumbnail findings into a concise CAP-style uterus pathology report.

Combine the visual findings into a concise final pathology report. State the specimen/procedure when supported, followed by the principal diagnosis and key qualifiers. Do not mention a reasoning process, chain of thought, answer keys, or unavailable prior diagnostic answers. If the thumbnail is insufficient for a diagnosis, say that the finding is insufficient or uncertain."""


def image_url(path: Path) -> str:
    mime = mimetypes.guess_type(path.name)[0] or "image/jpeg"
    data = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{data}"


def collect_images(paths: list[Path]) -> list[Path]:
    images: list[Path] = []
    for root in paths:
        if root.is_file() and root.suffix.lower() in IMAGE_SUFFIXES:
            images.append(root)
        elif root.is_dir():
            images.extend(p for p in root.iterdir() if p.suffix.lower() in IMAGE_SUFFIXES)
    return sorted(set(images), key=lambda p: p.name)


def build_content(image_path: Path) -> list[dict[str, Any]]:
    return [
        {"type": "text", "text": DIRECT_REPORT_PROMPT},
        {"type": "image_url", "image_url": {"url": image_url(image_path)}},
    ]


def open_workbook(output: Path) -> tuple[Workbook, Any]:
    if output.exists():
        wb = load_workbook(output)
        return wb, wb.active
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "thumbnail_direct_report"
    ws.append(["wsi_name", "vlm_result", "status", "thumbnail_path", "error", "created_at"])
    wb.save(output)
    return wb, ws


def load_done(output: Path) -> set[str]:
    if not output.exists():
        return set()
    wb = load_workbook(output, read_only=True)
    ws = wb.active
    done: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            done.add(str(row[0]))
    return done


def ask_vlm(
    client: OpenAI,
    *,
    model: str,
    image_path: Path,
    max_tokens: int,
) -> str:
    resp = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": build_content(image_path)},
        ],
        temperature=0.0,
        max_tokens=max_tokens,
    )
    return (resp.choices[0].message.content or "").strip()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--image-dir", type=Path, action="append", required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--endpoint", default="http://127.0.0.1:8000/v1")
    parser.add_argument("--model", default="/mnt/research/ljs/Qwen3-VL-8B-Instruct")
    parser.add_argument("--api-key", default="token-abc123")
    parser.add_argument("--limit", type=int, default=0, help="0 means all thumbnails")
    parser.add_argument("--max-tokens", type=int, default=256)
    parser.add_argument("--retries", type=int, default=2)
    parser.add_argument("--sleep", type=float, default=0.0)
    args = parser.parse_args()

    images = collect_images(args.image_dir)
    if args.limit:
        images = images[: args.limit]
    if not images:
        raise SystemExit("No thumbnail images found.")

    done = load_done(args.output)
    wb, ws = open_workbook(args.output)
    client = OpenAI(base_url=args.endpoint, api_key=args.api_key)

    print(f"Found {len(images)} thumbnails; already done {len(done)}; output={args.output}", flush=True)
    for idx, image_path in enumerate(images, start=1):
        wsi_name = image_path.stem
        if wsi_name in done:
            continue
        status = "ok"
        error = ""
        result = ""
        for attempt in range(1, args.retries + 1):
            try:
                result = ask_vlm(client, model=args.model, image_path=image_path, max_tokens=args.max_tokens)
                break
            except Exception as exc:
                status = "error"
                error = f"attempt_{attempt}: {type(exc).__name__}: {exc}"
                print(f"[{idx}/{len(images)}] {wsi_name} {error}", flush=True)
                time.sleep(min(30, 2**attempt))

        ws.append(
            [
                wsi_name,
                result,
                status if result else "error",
                str(image_path),
                error,
                datetime.now(timezone.utc).isoformat(),
            ]
        )
        wb.save(args.output)
        done.add(wsi_name)
        print(f"[{idx}/{len(images)}] {wsi_name} -> {status}: {result[:160]}", flush=True)
        if args.sleep:
            time.sleep(args.sleep)

    wb.save(args.output)
    print(f"Done: {args.output}", flush=True)


if __name__ == "__main__":
    main()
