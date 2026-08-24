#!/usr/bin/env python
from __future__ import annotations

import argparse
import json
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from agent.answers import normalize_answer
from agent.backends import ZeroShotQwenBackend
from agent.correction import check_consistency, should_retry
from agent.controller import build_query, chain_to_dict
from agent.memory import CaseMemory, JsonGraphStore
from agent.types import Step
from graph import GRAPH, ROOT_ID
from graph.schema import Node, NodeKind
from vision.backends import VisualBundle

try:
    from tqdm.auto import tqdm
except Exception:  # pragma: no cover - optional progress dependency
    tqdm = None


SAFE_LEVEL = {
    "1x": "1x",
    "1.25x": "1p25x",
    "2.5x": "2p5x",
    "5x": "5x",
    "10x": "10x",
    "20x": "20x",
    "40x": "40x",
}


def collect_slide_dirs(
    plip_root: Path,
    limit: int = 0,
    available_levels: list[str] | None = None,
) -> list[Path]:
    levels = available_levels or ["1x", "1.25x", "2.5x", "5x", "10x"]
    dirs: list[Path] = []
    for slide_dir in sorted((p for p in plip_root.iterdir() if p.is_dir()), key=lambda p: p.name):
        if any((slide_dir / SAFE_LEVEL[level] / "plip_topk_summary.json").exists() for level in levels):
            dirs.append(slide_dir)
            if limit and len(dirs) >= limit:
                break
    return dirs[:limit] if limit else dirs


def existing_slide_names(output: Path) -> set[str]:
    if not output.exists():
        return set()
    wb = load_workbook(output, read_only=True)
    ws = wb.active
    return {str(row[0]) for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0]}


def open_workbook(output: Path) -> tuple[Workbook, Any]:
    if output.exists():
        wb = load_workbook(output)
        return wb, wb.active
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "plip_patch_graph"
    ws.append(
        [
            "wsi_name",
            "vlm_result",
            "status",
            "node_path",
            "n_steps",
            "levels_used",
            "n_patch_images",
            "chain_json_path",
            "error",
            "created_at",
        ]
    )
    wb.save(output)
    return wb, ws


def load_summaries(slide_dir: Path, available_levels: list[str]) -> dict[str, dict[str, Any]]:
    summaries: dict[str, dict[str, Any]] = {}
    for level in available_levels:
        path = slide_dir / SAFE_LEVEL[level] / "plip_topk_summary.json"
        if not path.exists():
            continue
        try:
            summaries[level] = json.loads(path.read_text())
        except Exception:
            continue
    return summaries


def choose_level(node: Node, summaries: dict[str, dict[str, Any]], available_levels: list[str]) -> str | None:
    if not summaries:
        return None
    requested = node.mag_band
    requested_rank = _level_rank(requested)
    candidates = [level for level in available_levels if level in summaries]
    if requested in candidates:
        return requested
    lower_or_equal = [level for level in candidates if _level_rank(level) <= requested_rank]
    if lower_or_equal:
        return max(lower_or_equal, key=_level_rank)
    return max(candidates, key=_level_rank)


def _level_rank(level: str) -> float:
    return {
        "1x": 1.0,
        "1.25x": 1.25,
        "2.5x": 2.5,
        "5x": 5.0,
        "10x": 10.0,
        "20x": 20.0,
        "40x": 40.0,
    }.get(level, 0.0)


def find_thumbnail(slide_dir: Path, summaries: dict[str, dict[str, Any]], thumbnail_dir: Path | None) -> Path | None:
    for summary in summaries.values():
        value = summary.get("thumbnail_path")
        if value and Path(value).exists():
            return Path(value)
    if thumbnail_dir:
        stem = slide_dir.name
        candidates = [
            thumbnail_dir / f"{stem}.jpg",
            thumbnail_dir / f"{stem.removesuffix('.svs')}.jpg",
            thumbnail_dir / f"{stem.removesuffix('.svs')}.png",
        ]
        for path in candidates:
            if path.exists():
                return path
    return None


def patches_for_node(summary: dict[str, Any], node_id: str, top_k: int) -> list[Path]:
    for item in summary.get("nodes", []):
        if item.get("node_id") != node_id:
            continue
        paths: list[Path] = []
        for patch in item.get("top_patches", [])[:top_k]:
            path = Path(str(patch.get("image_path", "")))
            if path.exists():
                paths.append(path)
        return paths
    return []


def visual_for_node(
    node: Node,
    *,
    slide_dir: Path,
    summaries: dict[str, dict[str, Any]],
    thumbnail_path: Path | None,
    available_levels: list[str],
    top_k: int,
) -> tuple[VisualBundle, str, int]:
    level = choose_level(node, summaries, available_levels)
    patches: list[Path] = []
    if level is not None:
        patches = patches_for_node(summaries[level], node.id, top_k)
    metadata = {
        "visual": "thumbnail_plip_topk",
        "slide_id": slide_dir.name,
        "plip_level": level or "",
        "evidence_patch_count": len(patches),
        "embedding_context": (
            "UNI2/PLIP evidence selection was computed offline. "
            "Use the attached thumbnail and PLIP-retrieved patch images as visual evidence."
        ),
    }
    return (
        VisualBundle(
            thumbnail_path=thumbnail_path,
            patch_paths=patches,
            metadata=metadata,
        ),
        level or "",
        len(patches),
    )


def traverse_with_plip_visuals(
    backend: ZeroShotQwenBackend,
    *,
    slide_dir: Path,
    summaries: dict[str, dict[str, Any]],
    thumbnail_path: Path | None,
    available_levels: list[str],
    top_k: int,
    max_steps: int,
    confidence_threshold: float,
    max_answer_attempts: int,
) -> tuple[dict[str, Any], list[str], int]:
    store = JsonGraphStore(GRAPH)
    mem = CaseMemory()
    node = store.get_node(ROOT_ID)
    steps: list[Step] = []
    levels_used: list[str] = []
    n_patch_images = 0

    for _ in range(max_steps):
        query = build_query(node, steps)
        visual, level, patch_count = visual_for_node(
            node,
            slide_dir=slide_dir,
            summaries=summaries,
            thumbnail_path=thumbnail_path,
            available_levels=available_levels,
            top_k=top_k,
        )
        if level:
            levels_used.append(f"{node.id}:{level}")
        n_patch_images += patch_count
        extra = mem.retrieve_context(node, query)

        answer = ""
        confidence = 0.0
        last_raw = ""
        for attempt in range(max_answer_attempts):
            retry_note = ""
            if attempt:
                retry_note = (
                    "\nRetry instruction: your previous response was not a valid graph "
                    "answer. Return exactly one allowed answer key and nothing else."
                )
            raw, raw_confidence = backend.answer(
                node,
                visual,
                steps,
                extra_context=extra + retry_note,
            )
            last_raw = raw
            normalized = normalize_answer(raw, node)
            if normalized is None:
                continue
            if not answer or raw_confidence > confidence:
                answer, confidence = normalized, raw_confidence
            if not should_retry(confidence, confidence_threshold):
                break

        if not answer:
            raise ValueError(
                f"VLM failed to answer node {node.id!r}. Last response={last_raw!r}; "
                f"expected={node.options}"
            )

        _ = check_consistency(steps, node, answer)
        next_id = store.next(node.id, answer)
        next_question = store.get_node(next_id).question if next_id else ""
        step = Step(
            node.id,
            node.question,
            answer,
            confidence,
            next_question=next_question,
            raw_answer=last_raw,
        )
        steps.append(step)
        mem.append(node.id, node.question, answer)
        if next_id is None:
            break
        next_node = store.get_node(next_id)
        node = next_node
    else:
        raise RuntimeError(f"Traversal exceeded {max_steps} steps.")

    return chain_to_dict(steps, slide_id=slide_dir.name, include_report=True), levels_used, n_patch_images


def save_chain(chain: dict[str, Any], chains_dir: Path, slide_name: str) -> Path:
    chains_dir.mkdir(parents=True, exist_ok=True)
    path = chains_dir / f"{slide_name}.json"
    path.write_text(json.dumps(chain, indent=2, ensure_ascii=False) + "\n")
    return path


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--plip-root", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--chains-dir", type=Path, required=True)
    parser.add_argument("--thumbnail-dir", type=Path, default=Path("thumbnails"))
    parser.add_argument("--levels", default="1x,1.25x,2.5x,5x")
    parser.add_argument("--endpoint", default="http://127.0.0.1:8000/v1")
    parser.add_argument("--model", default="/mnt/research/ljs/Qwen3-VL-8B-Instruct")
    parser.add_argument("--api-key", default="token-abc123")
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--top-k", type=int, default=3)
    parser.add_argument("--max-steps", type=int, default=64)
    parser.add_argument("--confidence-threshold", type=float, default=0.65)
    parser.add_argument("--max-answer-attempts", type=int, default=2)
    parser.add_argument("--retries", type=int, default=2)
    args = parser.parse_args()

    import openai

    available_levels = [level.strip() for level in args.levels.split(",") if level.strip()]
    slide_dirs = collect_slide_dirs(args.plip_root, args.limit, available_levels)
    done = existing_slide_names(args.output)
    wb, ws = open_workbook(args.output)
    client = openai.OpenAI(base_url=args.endpoint, api_key=args.api_key)
    backend = ZeroShotQwenBackend(
        client,
        args.model,
        use_guided_choice=False,
        request_logprobs=False,
    )

    print(
        f"Found {len(slide_dirs)} slide directories; already done {len(done)}; output={args.output}",
        flush=True,
    )
    pending = [slide_dir for slide_dir in slide_dirs if slide_dir.name not in done]
    iterator = pending
    progress = None
    if tqdm is not None:
        progress = tqdm(
            pending,
            total=len(pending),
            desc="PLIP graph eval",
            unit="wsi",
            dynamic_ncols=True,
        )
        iterator = progress

    for idx, slide_dir in enumerate(iterator, start=1):
        slide_name = slide_dir.name
        if progress is not None:
            progress.set_postfix_str(slide_name[:48])
        status = "ok"
        error = ""
        chain: dict[str, Any] = {}
        levels_used: list[str] = []
        n_patch_images = 0
        try:
            summaries = load_summaries(slide_dir, available_levels)
            if not summaries:
                raise FileNotFoundError(f"No PLIP summaries found for {slide_name}")
            thumbnail_path = find_thumbnail(
                slide_dir,
                summaries,
                args.thumbnail_dir if args.thumbnail_dir.exists() else None,
            )
            for attempt in range(1, args.retries + 1):
                try:
                    chain, levels_used, n_patch_images = traverse_with_plip_visuals(
                        backend,
                        slide_dir=slide_dir,
                        summaries=summaries,
                        thumbnail_path=thumbnail_path,
                        available_levels=available_levels,
                        top_k=args.top_k,
                        max_steps=args.max_steps,
                        confidence_threshold=args.confidence_threshold,
                        max_answer_attempts=args.max_answer_attempts,
                    )
                    break
                except Exception as exc:
                    status = "error"
                    error = f"attempt_{attempt}: {type(exc).__name__}: {exc}"
                    print(f"[{idx}/{len(slide_dirs)}] {slide_name} {error}", flush=True)
                    time.sleep(min(30, 2**attempt))
        except Exception as exc:
            status = "error"
            error = f"{type(exc).__name__}: {exc}"

        report = str(chain.get("report") or "")
        node_path = " -> ".join(chain.get("node_path") or [])
        n_steps = len(chain.get("chain-of-thought") or [])
        chain_path = ""
        if chain:
            chain_path = str(save_chain(chain, args.chains_dir, slide_name))
            status = "ok"

        ws.append(
            [
                slide_name,
                report,
                status,
                node_path,
                n_steps,
                "; ".join(levels_used),
                n_patch_images,
                chain_path,
                error,
                datetime.now(timezone.utc).isoformat(),
            ]
        )
        wb.save(args.output)
        done.add(slide_name)
        message = f"[{idx}/{len(pending)}] {slide_name} -> {status}: {report[:160]}"
        if tqdm is not None:
            tqdm.write(message)
        else:
            print(message, flush=True)

    wb.save(args.output)
    print(f"Done: {args.output}", flush=True)


if __name__ == "__main__":
    main()
